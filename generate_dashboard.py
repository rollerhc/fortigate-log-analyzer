"""
generate_dashboard.py – Excel Dashboard Generator
Reads the JSON summary produced by fortianalyzer.py and exports
a formatted, multi-sheet Excel workbook with embedded charts.

Requirements: openpyxl
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, PieChart, Reference


# ===== PATHS =====
BASE_DIR     = Path(__file__).parent.parent
OUTPUT_DIR   = BASE_DIR / "output"
LAST_SUMMARY = OUTPUT_DIR / "last_summary.json"
PREV_SUMMARY = OUTPUT_DIR / "previous_summary.json"


# ============================================================
# STYLING HELPERS
# ============================================================

def auto_fit_columns(ws):
    """Widen each column to fit its longest value."""
    for col in ws.columns:
        max_len    = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2


def create_table(ws, table_name: str):
    """Wrap the worksheet data in a named Excel Table."""
    ref   = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_header(ws):
    """Apply a light-blue header style to the first row."""
    fill  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    font  = Font(bold=True, color="000000")
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align


def apply_borders(ws):
    """Add thin light-grey borders to every cell."""
    thin   = Side(border_style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


def freeze_header(ws):
    """Freeze the first row so it stays visible while scrolling."""
    ws.freeze_panes = "A2"


def format_numbers(ws, gb_cols=None, pct_cols=None):
    """Apply number formatting to specified columns."""
    if gb_cols:
        for col in gb_cols:
            for cell in ws[col]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"
    if pct_cols:
        for col in pct_cols:
            for cell in ws[col]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0%"


def full_style(ws, table_name: str, gb_cols=None, pct_cols=None):
    """Apply all styling in one call."""
    style_header(ws)
    create_table(ws, table_name)
    freeze_header(ws)
    apply_borders(ws)
    auto_fit_columns(ws)
    format_numbers(ws, gb_cols, pct_cols)


# ============================================================
# HELPERS
# ============================================================

def load_json(path: Path) -> dict | None:
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def to_gb(value: int | float) -> float:
    return value / (1024 ** 3)


# ============================================================
# SHEET BUILDERS
# ============================================================

def build_sheets(wb: Workbook, current: dict, previous: dict | None, metrics: dict):

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Indicator", "Value"])

    total_bytes = sum(d["bytes"] for d in current["dispositivos"].values())
    ws.append(["Total Consumption (GB)", round(to_gb(total_bytes), 2)])

    devs = sorted(current["dispositivos"].items(), key=lambda x: x[1]["bytes"], reverse=True)
    if devs:
        ip, info = devs[0]
        ws.append(["Top Consumer (IP)", ip])
        ws.append(["Top Consumer (GB)", round(to_gb(info["bytes"]), 2)])

    apps = sorted(current["apps"].items(), key=lambda x: x[1]["bytes"], reverse=True)
    if apps:
        app, info = apps[0]
        ws.append(["Top Application", app])
        ws.append(["Top App (GB)", round(to_gb(info["bytes"]), 2)])

    hours = sorted(current["horas"].items(), key=lambda x: x[1]["bytes"], reverse=True)
    if hours:
        h, info = hours[0]
        ws.append(["Peak Hour", h])
        ws.append(["Peak Consumption (GB)", round(to_gb(info["bytes"]), 2)])

    ws.append(["", ""])
    ws.append(["Category", "Percentage"])
    cats = sorted(metrics.get("por_categoria", []), key=lambda x: x["total_bytes"], reverse=True)
    for c in cats:
        ws.append([c["categoria"], c["percentual"] / 100])

    full_style(ws, "tblSummary", pct_cols=["B"])

    # ── Sheet 2: Top Devices ──────────────────────────────────
    ws = wb.create_sheet("Top Devices")
    ws.append(["IP", "Name", "MAC", "Total GB", "Sessions"])
    for ip, info in devs:
        m = next((d for d in metrics.get("por_dispositivo", []) if d["ip"] == ip), None)
        sessions = m["qtd_sessoes"] if m else 0
        ws.append([ip, info.get("srcname") or "", info.get("srcmac") or "",
                   round(to_gb(info["bytes"]), 3), sessions])
    full_style(ws, "tblTopDevices", gb_cols=["D"])

    # ── Sheet 3: Top Apps ─────────────────────────────────────
    ws = wb.create_sheet("Top Apps")
    ws.append(["Application", "Category", "Total GB", "Devices", "Sessions"])
    por_app = sorted(metrics.get("por_app", []), key=lambda x: x["total_bytes"], reverse=True)
    for item in por_app:
        ws.append([item["app"], item.get("category") or "",
                   round(item["total_gb"], 3), item["qtd_ips"], item["qtd_sessoes"]])
    full_style(ws, "tblTopApps", gb_cols=["C"])

    # ── Sheet 4: Categories ───────────────────────────────────
    ws = wb.create_sheet("Categories")
    ws.append(["Category", "Total GB", "%"])
    for c in cats:
        ws.append([c["categoria"], round(c["total_gb"], 3), c["percentual"] / 100])
    full_style(ws, "tblCategories", gb_cols=["B"], pct_cols=["C"])

    # ── Sheet 5: Internal vs External ────────────────────────
    ws = wb.create_sheet("Internal vs External")
    ws.append(["Type", "Total GB", "%"])
    for item in metrics.get("interno_externo", []):
        ws.append([item["tipo"], round(item["total_gb"], 3), item["percentual"] / 100])
    full_style(ws, "tblDirection", gb_cols=["B"], pct_cols=["C"])

    # ── Sheet 6: Hourly Consumption ───────────────────────────
    ws = wb.create_sheet("Hourly Consumption")
    ws.append(["Hour", "Total GB"])
    for item in sorted(metrics.get("consumo_por_hora", []), key=lambda x: x["hora"]):
        ws.append([item["hora"], round(item["total_gb"], 3)])
    full_style(ws, "tblHourly", gb_cols=["B"])

    # ── Sheet 7: Shadow IT ────────────────────────────────────
    ws = wb.create_sheet("Shadow IT")
    ws.append(["Application", "Total GB", "Devices"])
    shadow = sorted(metrics.get("shadow_it", []), key=lambda x: x["total_bytes"], reverse=True)
    for item in shadow:
        ws.append([item["app"], round(item["total_gb"], 3), item["qtd_ips"]])
    full_style(ws, "tblShadowIT", gb_cols=["B"])

    # ── Sheet 8: Frequency vs Volume ─────────────────────────
    ws = wb.create_sheet("Freq vs Volume")
    ws.append(["IP", "Total GB", "Sessions", "Avg GB/Session"])
    for item in metrics.get("por_dispositivo", []):
        gb_val   = item["total_gb"]
        sessions = item["qtd_sessoes"]
        avg      = gb_val / sessions if sessions > 0 else 0
        ws.append([item["ip"], round(gb_val, 3), sessions, round(avg, 4)])
    full_style(ws, "tblFreqVolume", gb_cols=["B", "D"])

    # ── Sheet 9: App Comparison (week-over-week) ──────────────
    ws = wb.create_sheet("App Comparison")
    ws.append(["Application", "Current GB", "Previous GB", "Delta GB"])
    if previous:
        apps_prev = previous.get("apps", {})
        for app, info in apps:
            cur_gb  = to_gb(info["bytes"])
            prev_gb = to_gb(apps_prev.get(app, {}).get("bytes", 0))
            delta   = cur_gb - prev_gb
            ws.append([app, round(cur_gb, 3), round(prev_gb, 3), round(delta, 3)])
    full_style(ws, "tblAppComparison", gb_cols=["B", "C", "D"])


# ============================================================
# CHART BUILDERS
# ============================================================

def build_charts(wb: Workbook):
    """Add one chart per relevant sheet."""

    def _bar(ws, title, y_title, x_title, data_col, cat_col, anchor, max_rows=16, data_col_end=None):
        max_r = min(ws.max_row, max_rows)
        chart = BarChart()
        chart.title        = title
        chart.y_axis.title = y_title
        chart.x_axis.title = x_title
        data = Reference(ws, min_col=data_col, max_col=data_col_end or data_col,
                         min_row=1, max_row=max_r)
        cats = Reference(ws, min_col=cat_col, min_row=2, max_row=max_r)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width  = 24
        ws.add_chart(chart, anchor)

    def _pie(ws, title, data_col, label_col, anchor):
        chart = PieChart()
        chart.title = title
        data   = Reference(ws, min_col=data_col, min_row=1, max_row=ws.max_row)
        labels = Reference(ws, min_col=label_col, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 12
        chart.width  = 18
        ws.add_chart(chart, anchor)

    # Top Devices – Bar
    _bar(wb["Top Devices"],   "Top Devices by Consumption (GB)", "GB", "IP",          4, 1, "H2")
    # Top Apps – Bar
    _bar(wb["Top Apps"],      "Top Applications by Consumption (GB)", "GB", "App",    3, 1, "H2")
    # Categories – Pie
    _pie(wb["Categories"],    "Traffic Categories", 2, 1, "E2")
    # Internal vs External – Pie
    _pie(wb["Internal vs External"], "Internal vs External", 2, 1, "E2")
    # Shadow IT – Bar
    _bar(wb["Shadow IT"],     "Shadow IT – Unrecognised Apps", "GB", "App",            2, 1, "E2")
    # Freq vs Volume – Bar (sessions)
    _bar(wb["Freq vs Volume"], "Top Offenders by Sessions", "Sessions", "IP",          3, 1, "F2")

    # Hourly Consumption – Line
    ws = wb["Hourly Consumption"]
    chart = LineChart()
    chart.title        = "Traffic by Hour"
    chart.y_axis.title = "GB"
    chart.x_axis.title = "Hour"
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width  = 24
    ws.add_chart(chart, "E2")

    # App Comparison – Grouped Bar
    ws = wb["App Comparison"]
    if ws.max_row > 1:
        max_r = min(ws.max_row, 16)
        chart = BarChart()
        chart.title        = "App Comparison (GB)"
        chart.y_axis.title = "GB"
        chart.x_axis.title = "Application"
        data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=max_r)
        cats = Reference(ws, min_col=1, min_row=2, max_row=max_r)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width  = 24
        ws.add_chart(chart, "F2")


# ============================================================
# ENTRY POINT
# ============================================================

def create_dashboard():
    current = load_json(LAST_SUMMARY)
    if not current:
        print("No last_summary.json found. Run fortianalyzer.py first.")
        return

    previous = load_json(PREV_SUMMARY)
    metrics  = current.get("metricas_avancadas", {})

    wb = Workbook()
    build_sheets(wb, current, previous, metrics)
    build_charts(wb)

    out_file = OUTPUT_DIR / "fortianalyzer_dashboard.xlsx"
    wb.save(out_file)
    print(f"Dashboard saved to: {out_file}")


if __name__ == "__main__":
    create_dashboard()
